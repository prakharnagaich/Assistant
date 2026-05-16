"""
NSE Option Chain Automation Application

This application automates:
1. Downloading CSV from NSE India option-chain page
2. Identifying and clicking download link based on image reference
3. Waiting for file completion
4. Copying downloaded sheet content to prediction.xlsx RAW sheet
5. Cleaning up downloaded file

Usage:
    python nse_automation_app.py --download-dir C:\Users\Prakhar\Desktop\AICode\downloads --interval 120

Requirements:
    pip install selenium webdriver-manager pandas openpyxl pillow opencv-python
"""

import time
import sys
import logging
from pathlib import Path
from datetime import datetime
from typing import Optional, Tuple

import pandas as pd
from openpyxl import load_workbook, Workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    TimeoutException,
    NoSuchElementException,
)
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('nse_automation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# Constants
NSE_URL = "https://www.nseindia.com/option-chain"
DEFAULT_BASE_DIR = r"C:\Users\Prakhar\Desktop\AICode"
DEFAULT_DOWNLOADS_DIR = Path(DEFAULT_BASE_DIR) / "downloads"
DEFAULT_PREDICTION_FILE = Path(DEFAULT_BASE_DIR) / "prediction.xlsx"
DEFAULT_IMAGE_NAME = "abcd.png"
DOWNLOAD_TIMEOUT = 180  # seconds
PAGE_LOAD_WAIT = 5  # seconds


class NSEAutomationApp:
    """Main automation application for NSE option-chain download and paste operations."""

    def __init__(
        self,
        download_dir: str = str(DEFAULT_DOWNLOADS_DIR),
        image_name: str = DEFAULT_IMAGE_NAME,
        prediction_file: str = str(DEFAULT_PREDICTION_FILE),
        headless: bool = True,
        base_dir: str = DEFAULT_BASE_DIR,
    ):
        """
        Initialize the automation application.

        Args:
            download_dir: Directory to save downloads
            image_name: Name of image file to click (e.g., 'abcd.png')
            prediction_file: Path to prediction.xlsx file
            headless: Run browser in headless mode
            base_dir: Base directory where image and prediction files are located
        """
        self.download_dir = Path(download_dir)
        self.image_name = image_name
        self.image_path = Path(base_dir) / image_name
        self.prediction_file = Path(prediction_file)
        self.headless = headless
        self.driver = None

        # Create downloads directory if it doesn't exist
        self.download_dir.mkdir(parents=True, exist_ok=True)

        logger.info(f"Initialized NSEAutomationApp")
        logger.info(f"  Download dir: {self.download_dir}")
        logger.info(f"  Image file: {self.image_path}")
        logger.info(f"  Prediction file: {self.prediction_file}")

    def _create_driver(self) -> webdriver.Chrome:
        """
        Create and configure Chrome WebDriver.

        Returns:
            Configured Chrome WebDriver instance
        """
        download_dir_str = str(self.download_dir.resolve())
        options = webdriver.ChromeOptions()

        # User agent to reduce detection
        options.add_argument(
            "--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )

        # Download preferences
        prefs = {
            "download.default_directory": download_dir_str,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True,
        }
        options.add_experimental_option("prefs", prefs)

        # Stability options
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-infobars")
        options.add_argument("--disable-blink-features=AutomationControlled")

        if self.headless:
            options.add_argument("--headless=new")

        # Create driver
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        driver.set_page_load_timeout(60)

        logger.info("Chrome WebDriver created successfully")
        return driver

    def _find_and_click_download_link(self) -> bool:
        """
        Find and click the download link based on image reference.

        Returns:
            True if link was clicked, False otherwise
        """
        logger.info(f"Searching for download link using image: {self.image_name}")

        wait = WebDriverWait(self.driver, 30)

        # Wait for page to load
        try:
            wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            logger.info("Page body loaded")
        except TimeoutException:
            logger.warning("Timeout waiting for page body")
            pass

        # Dismiss any overlays/cookies
        self._dismiss_overlays()

        # XPath strategies to find the download link
        stem = Path(self.image_name).stem.lower()
        xpaths = [
            # Priority 1: Anchor containing image with matching filename
            f"//a[img[contains(@src, '{self.image_name}')]]",
            f"//a[img[contains(@src, '{stem}')]]",
            # Priority 2: Anchor with style referencing image
            f"//a[contains(@style, '{self.image_name}')]",
            f"//a[contains(@style, '{stem}')]",
            f"//a[descendant::*[contains(@style, '{self.image_name}')]]",
            # Priority 3: Direct image element
            f"//img[contains(@src, '{self.image_name}')]",
            f"//img[contains(@src, '{stem}')]",
            f"//img[contains(translate(@alt, 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{stem}')]",
            # Priority 4: Any element with style referencing image
            f"//*[contains(@style, '{self.image_name}')]",
        ]

        for idx, xp in enumerate(xpaths, 1):
            try:
                logger.info(f"Trying XPath strategy {idx}")
                elems = self.driver.find_elements(By.XPATH, xp)

                if not elems:
                    continue

                for el in elems:
                    if not el.is_displayed():
                        logger.debug(f"Element not displayed, skipping")
                        continue

                    # Prefer ancestor anchor for clicks
                    target = el
                    if el.tag_name.lower() != "a":
                        try:
                            parent_a = el.find_element(By.XPATH, "./ancestor::a[1]")
                            if parent_a and parent_a.is_displayed():
                                target = parent_a
                                logger.info("Found parent anchor element")
                        except NoSuchElementException:
                            pass

                    # Click the element
                    logger.info(f"Scrolling element into view")
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", target)
                    time.sleep(0.5)

                    try:
                        logger.info(f"Attempting to click element")
                        target.click()
                        logger.info(f"Successfully clicked element")
                        return True
                    except ElementClickInterceptedException:
                        logger.info(f"Normal click intercepted, using JS click")
                        try:
                            self.driver.execute_script("arguments[0].click();", target)
                            logger.info(f"Successfully JS-clicked element")
                            return True
                        except Exception as e:
                            logger.debug(f"JS click failed: {e}")
                            continue

            except Exception as e:
                logger.debug(f"XPath strategy {idx} failed: {e}")
                continue

        logger.error("Could not find or click download link")
        return False

    def _dismiss_overlays(self) -> None:
        """Dismiss any overlay dialogs (cookies, popups, etc.)."""
        logger.info("Attempting to dismiss overlays...")
        overlay_xpaths = [
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'accept')]",
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'agree')]",
            "//button[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'ok')]",
            "//a[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'accept')]",
        ]

        for xp in overlay_xpaths:
            try:
                elems = self.driver.find_elements(By.XPATH, xp)
                for el in elems:
                    if el.is_displayed():
                        try:
                            el.click()
                            logger.info("Dismissed overlay")
                            time.sleep(1)
                        except Exception:
                            try:
                                self.driver.execute_script("arguments[0].click();", el)
                                logger.info("Dismissed overlay via JS")
                            except Exception:
                                pass
            except Exception:
                pass

    def _wait_for_download(self) -> Path:
        """
        Wait for a file to be downloaded to the download directory.

        Returns:
            Path to the downloaded file

        Raises:
            TimeoutError: If no file appears within timeout period
        """
        logger.info(f"Waiting for file download (timeout: {DOWNLOAD_TIMEOUT}s)...")

        start_time = time.time()
        seen_files = set(self.download_dir.iterdir()) if self.download_dir.exists() else set()

        while time.time() - start_time < DOWNLOAD_TIMEOUT:
            time.sleep(1)

            if not self.download_dir.exists():
                continue

            current_files = set(self.download_dir.iterdir())
            new_files = current_files - seen_files

            # Filter out incomplete downloads
            candidates = [
                f for f in new_files
                if not f.name.endswith('.crdownload')
                and not f.name.endswith('.part')
                and not f.name.endswith('.tmp')
            ]

            if candidates:
                # Sort by modification time, get newest
                candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
                downloaded_file = candidates[0]

                logger.info(f"New file detected: {downloaded_file.name}")

                # Wait for file size to stabilize
                logger.info("Waiting for file to stabilize...")
                prev_size = -1
                stable_count = 0

                while time.time() - start_time < DOWNLOAD_TIMEOUT:
                    current_size = downloaded_file.stat().st_size

                    if current_size == prev_size:
                        stable_count += 1
                        if stable_count >= 3:
                            logger.info(f"File download complete: {downloaded_file} ({current_size} bytes)")
                            return downloaded_file
                    else:
                        prev_size = current_size
                        stable_count = 0
                        logger.debug(f"File size: {current_size} bytes")

                    time.sleep(0.5)

            # Check existing files for completed downloads
            for f in current_files:
                if f.name.endswith('.crdownload'):
                    continue
                if time.time() - f.stat().st_mtime > 1:
                    logger.info(f"Found existing file: {f.name}")
                    return f

        raise TimeoutError(f"No file downloaded within {DOWNLOAD_TIMEOUT}s")

    def _read_downloaded_file(self, file_path: Path) -> pd.DataFrame:
        """
        Read downloaded file as DataFrame.

        Args:
            file_path: Path to downloaded file

        Returns:
            DataFrame with file contents

        Raises:
            RuntimeError: If file cannot be read
        """
        logger.info(f"Reading downloaded file: {file_path}")

        try:
            # Try reading as Excel first
            try:
                df = pd.read_excel(file_path)
                logger.info(f"Successfully read file as Excel")
            except Exception as e:
                logger.warning(f"Excel read failed: {e}, trying CSV...")
                df = pd.read_csv(file_path)
                logger.info(f"Successfully read file as CSV")

            logger.info(f"Data shape: {df.shape}")
            return df

        except Exception as e:
            logger.error(f"Failed to read file: {e}")
            raise RuntimeError(f"Could not read downloaded file: {e}")

    def _paste_to_prediction(self, df: pd.DataFrame) -> None:
        """
        Paste DataFrame content to prediction.xlsx RAW sheet.

        Args:
            df: DataFrame to paste

        Raises:
            RuntimeError: If paste operation fails
        """
        logger.info(f"Pasting data to prediction file: {self.prediction_file}")

        try:
            # Load or create prediction workbook
            if self.prediction_file.exists():
                logger.info("Loading existing prediction workbook...")
                book = load_workbook(self.prediction_file)
            else:
                logger.info("Creating new prediction workbook...")
                book = Workbook()
                # Remove default sheet
                if 'Sheet' in book.sheetnames:
                    book.remove(book['Sheet'])

            # Remove existing RAW sheet if present
            if 'RAW' in book.sheetnames:
                logger.info("Removing existing RAW sheet...")
                book.remove(book['RAW'])

            # Use pandas to write data
            logger.info("Writing data to RAW sheet...")
            with pd.ExcelWriter(self.prediction_file, engine='openpyxl', mode='w') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name='RAW', index=False)

            logger.info(f"Successfully pasted data to RAW sheet")
            logger.info(f"Data pasted: {len(df)} rows, {len(df.columns)} columns")

        except Exception as e:
            logger.error(f"Failed to paste data: {e}")
            raise RuntimeError(f"Could not paste to prediction file: {e}")

    def _cleanup_downloaded_file(self, file_path: Path) -> None:
        """
        Delete the original downloaded file.

        Args:
            file_path: Path to file to delete
        """
        logger.info(f"Cleaning up downloaded file: {file_path}")

        try:
            file_path.unlink()
            logger.info(f"Successfully deleted file")
        except Exception as e:
            logger.warning(f"Failed to delete file: {e}")

    def run_once(self) -> bool:
        """
        Execute one complete automation cycle.

        Returns:
            True if successful, False otherwise
        """
        logger.info("=" * 80)
        logger.info(f"Starting automation cycle at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        logger.info("=" * 80)

        try:
            # Create driver
            self.driver = self._create_driver()

            # Open NSE URL
            logger.info(f"Opening NSE URL: {NSE_URL}")
            self.driver.get(NSE_URL)
            time.sleep(PAGE_LOAD_WAIT)

            # Find and click download link
            if not self._find_and_click_download_link():
                logger.error("Failed to click download link")
                return False

            # Wait for download
            downloaded_file = self._wait_for_download()

            # Read downloaded file
            df = self._read_downloaded_file(downloaded_file)

            # Paste to prediction
            self._paste_to_prediction(df)

            # Cleanup
            self._cleanup_downloaded_file(downloaded_file)

            logger.info("=" * 80)
            logger.info(f"Automation cycle completed successfully")
            logger.info("=" * 80)
            return True

        except Exception as e:
            logger.error(f"Automation cycle failed: {e}", exc_info=True)
            logger.info("=" * 80)
            return False

        finally:
            # Close driver
            if self.driver:
                try:
                    self.driver.quit()
                    logger.info("Browser closed")
                except Exception as e:
                    logger.warning(f"Error closing browser: {e}")

    def run_continuous(self, interval: int = 120) -> None:
        """
        Run automation continuously at specified intervals.

        Args:
            interval: Interval between runs in seconds (default: 120)
        """
        logger.info(f"Starting continuous automation (interval: {interval}s)")
        logger.info("Press Ctrl+C to stop")

        try:
            while True:
                success = self.run_once()
                if success:
                    logger.info(f"Sleeping for {interval} seconds until next run...")
                else:
                    logger.warning(f"Run failed, retrying in {interval} seconds...")

                time.sleep(interval-60)

        except KeyboardInterrupt:
            logger.info("Automation stopped by user")
            sys.exit(0)
        except Exception as e:
            logger.error(f"Fatal error: {e}", exc_info=True)
            sys.exit(1)


def main():
    """Main entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="NSE Option Chain Automation Application",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Run once
  python nse_automation_app.py --run-once

  # Run continuously every 2 minutes
  python nse_automation_app.py --interval 120

  # Custom directories
  python nse_automation_app.py --download-dir C:\\Downloads --prediction C:\\files\\pred.xlsx
        """
    )

    parser.add_argument(
        "--download-dir",
        type=str,
        default=str(DEFAULT_DOWNLOADS_DIR),
        help=f"Directory to save downloads (default: {DEFAULT_DOWNLOADS_DIR})"
    )
    parser.add_argument(
        "--image",
        type=str,
        default=DEFAULT_IMAGE_NAME,
        help=f"Image file name to find (default: {DEFAULT_IMAGE_NAME})"
    )
    parser.add_argument(
        "--prediction",
        type=str,
        default=str(DEFAULT_PREDICTION_FILE),
        help=f"Path to prediction.xlsx (default: {DEFAULT_PREDICTION_FILE})"
    )
    parser.add_argument(
        "--base-dir",
        type=str,
        default=DEFAULT_BASE_DIR,
        help=f"Base directory for image and prediction files (default: {DEFAULT_BASE_DIR})"
    )
    parser.add_argument(
        "--interval",
        type=int,
        default=120,
        help="Interval for continuous runs in seconds (default: 120)"
    )
    parser.add_argument(
        "--run-once",
        action="store_true",
        help="Run automation only once instead of continuous"
    )
    parser.add_argument(
        "--no-headless",
        action="store_true",
        help="Run browser in non-headless mode (useful for debugging)"
    )

    args = parser.parse_args()

    # Create application
    app = NSEAutomationApp(
        download_dir=args.download_dir,
        image_name=args.image,
        prediction_file=args.prediction,
        headless=not args.no_headless,
        base_dir=args.base_dir,
    )

    # Run
    if args.run_once:
        success = app.run_once()
        sys.exit(0 if success else 1)
    else:
        app.run_continuous(interval=args.interval)


if __name__ == "__main__":
    main()
