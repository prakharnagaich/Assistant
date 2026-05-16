"""
NSE Option Chain Automation Script  (v2)
=========================================
Changes from v1:
  • Downloaded CSV is always renamed  →  abcd.csv
  • Before each cycle starts, any existing abcd.csv is deleted first
  • On success the ENTIRE prediction.xlsx is refreshed (all sheets cleared &
    re-written with the new CSV data in the first sheet)
  • Runs every 2 minutes continuously

Directory : C:\\Users\\Prakhar\\Desktop\\AICode
Files needed in that directory:
  abcd.png        – screenshot of the download button (for image-match fallback)
  prediction.xlsx – workbook to refresh
"""

import os
import sys
import time
import glob
import logging
import traceback
from datetime import datetime

import cv2
import numpy as np
import pyautogui
import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


# ══════════════════════════════════════════════════════════════
#  CONFIGURATION  –  edit only this block
# ══════════════════════════════════════════════════════════════
BASE_DIR         = r"C:\Users\Prakhar\Desktop\AICode"
DOWNLOAD_DIR     = BASE_DIR
PREDICTION_FILE  = os.path.join(BASE_DIR, "prediction.xlsx")
REFERENCE_IMG    = os.path.join(BASE_DIR, "abcd.png")     # button screenshot
DOWNLOADED_CSV   = os.path.join(BASE_DIR, "abcd.csv")     # fixed target name
NSE_URL          = "https://www.nseindia.com/option-chain"
INTERVAL_SEC     = 120    # 2 minutes between cycles
DOWNLOAD_WAIT    = 60     # max seconds to wait for Chrome to finish the download
IMG_CONFIDENCE   = 0.70   # OpenCV match threshold  (0 – 1)
# ══════════════════════════════════════════════════════════════


# ──────────────────────────────────────────────────────────────
#  LOGGING  (file + console)
# ──────────────────────────────────────────────────────────────
_log_path = os.path.join(BASE_DIR, "nse_automation.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)s]  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(_log_path, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("nse_auto")


# ══════════════════════════════════════════════════════════════
#  STEP 1 – BROWSER SETUP
# ══════════════════════════════════════════════════════════════
def build_driver() -> webdriver.Chrome:
    """
    Create a Chrome WebDriver that:
      • downloads files silently to DOWNLOAD_DIR
      • mimics a real browser (UA + maximised window)
    """
    opts = Options()
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-notifications")
    opts.add_argument("--disable-popup-blocking")
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
    opts.add_experimental_option("prefs", {
        "download.default_directory":   DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade":   True,
        "safebrowsing.enabled":         True,
    })
    svc = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=svc, options=opts)


def dismiss_cookie_banner(driver: webdriver.Chrome, timeout: int = 5) -> None:
    """Click Accept/Agree on any cookie overlay NSE may show."""
    try:
        btn = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH,
                "//button["
                "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'accept')"
                " or "
                "contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'agree')"
                "]"
            ))
        )
        btn.click()
        log.info("Cookie banner dismissed.")
    except Exception:
        pass   # no banner – that is fine


# ══════════════════════════════════════════════════════════════
#  STEP 2a – CLICK DOWNLOAD  via XPath / CSS selectors
# ══════════════════════════════════════════════════════════════
_SELECTORS = [
    (By.CSS_SELECTOR, "a.downloadCSV"),
    (By.CSS_SELECTOR, "a[href*='.csv']"),
    (By.XPATH,  "//a[contains(@href,'csv') or contains(@href,'CSV')]"),
    (By.XPATH,  "//a[contains(@title,'Download') or contains(@title,'download')]"),
    (By.XPATH,  "//img[contains(@src,'download') or contains(@alt,'download')]/ancestor::a"),
    (By.XPATH,  "//img[contains(@src,'csv') or contains(@alt,'csv')]/ancestor::a"),
    (By.XPATH,  "//span[contains(@class,'download')]/ancestor::a"),
    (By.XPATH,  "//a[contains(@class,'download') or contains(@class,'Download')]"),
    (By.XPATH,  "//*[@id='downloadOCTable']"),
    (By.XPATH,  "//*[contains(@id,'download') or contains(@id,'Download')]"),
]


def click_via_selectors(driver: webdriver.Chrome, timeout: int = 15) -> bool:
    """Return True if one selector matched and the element was clicked."""
    for by, sel in _SELECTORS:
        try:
            el = WebDriverWait(driver, timeout).until(
                EC.element_to_be_clickable((by, sel))
            )
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.4)
            el.click()
            log.info("Clicked via selector  [%s]  %s", by, sel)
            return True
        except Exception:
            continue
    log.warning("All %d selectors failed.", len(_SELECTORS))
    return False


# ══════════════════════════════════════════════════════════════
#  STEP 2b – CLICK DOWNLOAD  via OpenCV image recognition
# ══════════════════════════════════════════════════════════════
def click_via_image(ref_path: str, confidence: float = IMG_CONFIDENCE) -> bool:
    """
    Screenshot the entire screen, run template matching against *ref_path*,
    and click the centre of the best match if confidence is met.
    """
    if not os.path.exists(ref_path):
        log.warning("Reference image missing: %s", ref_path)
        return False

    template = cv2.imread(ref_path, cv2.IMREAD_COLOR)
    if template is None:
        log.warning("cv2 could not read reference image: %s", ref_path)
        return False

    screen    = pyautogui.screenshot()
    screen_np = cv2.cvtColor(np.array(screen), cv2.COLOR_RGB2BGR)
    result    = cv2.matchTemplate(screen_np, template, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, max_loc = cv2.minMaxLoc(result)

    log.info("Image-match confidence = %.3f  (threshold = %.2f)", max_val, confidence)
    if max_val < confidence:
        log.warning("Image match below threshold – download button not found on screen.")
        return False

    th, tw  = template.shape[:2]
    cx, cy  = max_loc[0] + tw // 2, max_loc[1] + th // 2
    pyautogui.click(cx, cy)
    log.info("Clicked via image recognition at screen coords (%d, %d)", cx, cy)
    return True


# ══════════════════════════════════════════════════════════════
#  STEP 3 – WAIT FOR DOWNLOAD + RENAME TO abcd.csv
# ══════════════════════════════════════════════════════════════
def wait_and_rename(before_snapshot: set, timeout: int = DOWNLOAD_WAIT) -> bool:
    """
    Poll DOWNLOAD_DIR until a new *.csv file appears (Chrome's .crdownload
    must be gone and file size must stabilise).  Then rename to abcd.csv.

    Returns True on success, False on timeout.
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        partials = glob.glob(os.path.join(DOWNLOAD_DIR, "*.crdownload"))
        if partials:
            log.debug("Download in progress (%d .crdownload file(s))…", len(partials))
            time.sleep(2)
            continue

        current  = set(glob.glob(os.path.join(DOWNLOAD_DIR, "*.csv")))
        new_csvs = current - before_snapshot
        if new_csvs:
            raw_path = sorted(new_csvs, key=os.path.getmtime)[-1]

            # Wait until file size stabilises (written fully)
            prev_size    = -1
            stable_ticks = 0
            while stable_ticks < 2:
                size = os.path.getsize(raw_path)
                stable_ticks = (stable_ticks + 1) if (size > 0 and size == prev_size) else 0
                prev_size    = size
                time.sleep(1)

            # Delete previous abcd.csv (if exists) then rename new download → abcd.csv
            if os.path.exists(DOWNLOADED_CSV):
                os.remove(DOWNLOADED_CSV)
                log.info("Deleted previous abcd.csv before renaming new download.")
            os.rename(raw_path, DOWNLOADED_CSV)
            log.info("File saved as: %s  (%d bytes)",
                     DOWNLOADED_CSV, os.path.getsize(DOWNLOADED_CSV))
            return True

        time.sleep(2)

    log.error("Timed out after %d s waiting for CSV download.", timeout)
    return False


# ══════════════════════════════════════════════════════════════
#  STEP 4 – REFRESH ENTIRE prediction.xlsx
# ══════════════════════════════════════════════════════════════
def refresh_prediction_xlsx() -> None:
    """
    Fully refresh prediction.xlsx with the contents of abcd.csv:
      1. Read abcd.csv into a DataFrame.
      2. Open prediction.xlsx.
      3. Clear ALL sheets completely (every row in every sheet).
      4. Write the CSV data (header + all rows) into the FIRST sheet.
      5. Save and close.
    """
    log.info("Reading abcd.csv…")
    df = pd.read_csv(DOWNLOADED_CSV, dtype=str).fillna("")

    log.info("Opening prediction.xlsx…")
    wb = openpyxl.load_workbook(PREDICTION_FILE)

    # ── Clear every sheet ────────────────────────────────
    for name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row + 1)
        log.info("  Cleared sheet: '%s'", name)

    # ── Write CSV data into first sheet ──────────────────
    first_ws = wb.worksheets[0]
    first_ws.append(list(df.columns))          # header
    for _, row in df.iterrows():
        first_ws.append(list(row))             # data

    wb.save(PREDICTION_FILE)
    log.info("prediction.xlsx refreshed: %d data rows → sheet '%s'",
             len(df), first_ws.title)


# ══════════════════════════════════════════════════════════════
#  MAIN CYCLE  (one complete run)
# ══════════════════════════════════════════════════════════════
def run_cycle(cycle_num: int) -> None:
    """
    Pipeline:
      1  Open NSE option-chain page
      2  Click the CSV download button
      3  Wait for download → delete old abcd.csv → rename new file to abcd.csv
      4  Refresh prediction.xlsx
         (abcd.csv is kept in the directory after every cycle)
    """
    log.info("━" * 64)
    log.info("CYCLE #%d  |  %s", cycle_num,
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    driver = build_driver()
    try:
        # 1 – open page
        log.info("Opening: %s", NSE_URL)
        driver.get(NSE_URL)
        dismiss_cookie_banner(driver)

        log.info("Waiting for option-chain table to render…")
        try:
            WebDriverWait(driver, 40).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table, .ocTable"))
            )
            log.info("Table visible – page loaded.")
        except Exception:
            log.warning("Table not detected in 40 s – continuing anyway.")

        time.sleep(4)  # extra JS-settle buffer

        # Snapshot current CSVs before triggering download
        before = set(glob.glob(os.path.join(DOWNLOAD_DIR, "*.csv")))

        # 2 – click download (selector first, image fallback)
        clicked = click_via_selectors(driver)
        if not clicked:
            log.info("Selector approach failed – switching to image recognition…")
            clicked = click_via_image(REFERENCE_IMG)
        if not clicked:
            raise RuntimeError(
                "Could not locate or click the CSV download button by any method."
            )

        # 3 – wait for file, rename to abcd.csv
        if not wait_and_rename(before):
            raise RuntimeError("abcd.csv was not created within the timeout window.")

        # 4 – refresh prediction.xlsx
        refresh_prediction_xlsx()

        log.info("CYCLE #%d  COMPLETE  (abcd.csv retained in directory)", cycle_num)

    except Exception as exc:
        log.error("CYCLE #%d  FAILED: %s", cycle_num, exc)
        log.debug(traceback.format_exc())
    finally:
        try:
            driver.quit()
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════
#  STARTUP VALIDATION
# ══════════════════════════════════════════════════════════════
def validate() -> None:
    errors = []
    if not os.path.isdir(BASE_DIR):
        errors.append(f"BASE_DIR not found       : {BASE_DIR}")
    if not os.path.isfile(PREDICTION_FILE):
        errors.append(f"prediction.xlsx not found: {PREDICTION_FILE}")
    if not os.path.isfile(REFERENCE_IMG):
        log.warning("abcd.png not found – image-recognition fallback DISABLED.")
    if errors:
        for e in errors:
            log.error("STARTUP ERROR: %s", e)
        raise SystemExit("Resolve the above errors and restart.")


# ══════════════════════════════════════════════════════════════
#  ENTRY POINT  –  infinite loop, every INTERVAL_SEC seconds
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    validate()

    log.info("╔══════════════════════════════════════════════╗")
    log.info("║   NSE OPTION CHAIN AUTOMATION  v2           ║")
    log.info("╠══════════════════════════════════════════════╣")
    log.info("║  Base dir   : %-29s ║", BASE_DIR[-28:])
    log.info("║  CSV name   : abcd.csv                      ║")
    log.info("║  Workbook   : prediction.xlsx               ║")
    log.info("║  Interval   : %-29s ║", f"every {INTERVAL_SEC}s")
    log.info("╚══════════════════════════════════════════════╝")
    log.info("Press Ctrl+C to stop.")

    cycle_count = 0
    while True:
        cycle_count += 1
        run_cycle(cycle_count)
        log.info("Sleeping %d s before next cycle…", INTERVAL_SEC)
        time.sleep(INTERVAL_SEC)
